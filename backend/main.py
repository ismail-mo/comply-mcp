"""COMPLY API — FastAPI backend.

Architecture (the deterministic reframe, spec Part 1):
- A compliance audit is a FIXED pipeline, not an agentic loop:
    Stage 0  upload -> parse -> document_index          (deterministic)
    Stage 1  route intent                                (keyword, deterministic)
    Stage 2  identify element via the index              (deterministic)
    Stage 4  extract_element                             (AI call #1, schema-enforced)
    Stage 5  run_audit engine                            (pure Python, <100 ms)
    Stage 6  write_overview                              (AI call #2, narration only)
    Stage 7  assemble + persist + stream JSON            (deterministic)
- TWO AI calls per element. The AI never runs a formula and never decides a
  verdict — formulas/, tables/, registries/, engine/ at the repo root own all
  engineering, gated by known-answer tests.
- quick_answer path: retrieval (search_documents) + one streamed prose call.
- Token discipline: tools return compact structured JSON; the orchestrator
  context stays flat; history is trimmed server-side.
"""

import asyncio
import json
import logging
import os
import re
import sys
import time
import uuid
from contextlib import asynccontextmanager
from datetime import datetime, timezone
from pathlib import Path

import anthropic
from dotenv import load_dotenv
from fastapi import FastAPI, File, Header, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import load_workbook
from pydantic import BaseModel, Field
from typing import Literal, Optional

# Repo-root packages: engine/, formulas/, tables/, registries/
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from mcp_client import MCPClient
from file_reader import read_file, read_pdf_sample, extract_pdf_coordinates
from services.indexer import build_document_index

from engine.run_audit import run_audit, summarize
from engine.cross_reference import cross_reference
from registries.check_registry import CHECK_REGISTRY
from tables.section_properties import get_section, normalize_designation

load_dotenv()

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("comply.main")

ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")
UPLOAD_DIR = Path(os.getenv("UPLOAD_DIR", "./uploads"))
AUDITS_DIR = Path(os.getenv("AUDITS_DIR", "./audits"))
ANTHROPIC_MODEL = os.getenv("ANTHROPIC_MODEL", "claude-sonnet-4-6")
ALLOWED_ORIGINS = [
    origin.strip()
    for origin in os.getenv(
        "CORS_ORIGINS",
        "http://localhost:3000,http://localhost:3001,http://localhost:3002",
    ).split(",")
    if origin.strip()
]
MAX_CONCURRENT_CHATS = int(os.getenv("MAX_CONCURRENT_CHATS", "3"))
ADMIN_API_KEY = os.getenv("ADMIN_API_KEY")

MAX_HISTORY_MESSAGES = 8
MAX_HISTORY_CHARS = 4000

anthropic_client = anthropic.AsyncAnthropic(api_key=ANTHROPIC_API_KEY)

ALLOWED_EXTENSIONS = {".pdf", ".xlsx", ".xls", ".csv"}

_coords_cache: dict = {}
_COORDS_CACHE_MAX = 8
_request_lock = asyncio.Semaphore(MAX_CONCURRENT_CHATS)

COMPLIANCE_AUDIT_PREFIX = "[COMPLIANCE_AUDIT]"

# Extracted names the AI must look for, per element type, with the exact
# semantics — descriptions travel into the extraction prompt so the model
# cannot mis-map (e.g. a required AREA is not a compression RESISTANCE).
INPUT_DESCRIPTIONS = {
    "NEd": "TOTAL design (factored) axial load on the element, kN",
    "VEd": "design shear force, kN",
    "MEd": "maximum design bending moment, kNm",
    "Gk": "TOTAL UNFACTORED permanent (dead) action for the element, kN — sum across floors if tabulated; NEVER a factored value",
    "Qk": "TOTAL UNFACTORED variable (imposed) action for the element, kN — sum across floors if tabulated; NEVER a factored value",
    "Le": "effective / buckling length of the column, with the unit as stated (m or mm)",
    "fy": "steel yield strength used in the calculations, N/mm2",
    "grade": "steel grade, e.g. S275 or S355",
    "section_class": "cross-section class the document claims (1-4)",
    "gamma_Q_used": "the load factor the document actually applies to VARIABLE/IMPOSED loads in its combination (e.g. 1.35 or 1.5)",
    "compression_result": "the compression RESISTANCE Nc,Rd in kN that the document claims — NOT a required area, NOT a demand",
    "flex_buckling_result": "the member buckling RESISTANCE Nb,Rd in kN that the document claims (e.g. from Blue Book tables)",
    "shear_result": "the shear RESISTANCE Vpl,Rd in kN that the document claims",
    "bending_result": "the bending RESISTANCE Mc,Rd in kNm that the document claims",
    "ltb_result": "the buckling moment RESISTANCE Mb,Rd in kNm that the document claims (e.g. from Blue Book tables)",
    "deflection_result": "the SLS deflection in mm that the document claims",
    "q": "SLS UDL on the beam from VARIABLE actions, kN/m",
    "P": "SLS midspan point load from VARIABLE actions, kN",
    "L": "beam span, with the unit as stated (m or mm)",
    "Mcr": "elastic critical LTB moment if stated, kNm",
    "C1": "moment-distribution factor C1 if stated",
}

REQUIRED_EXTRACT_INPUTS = {
    "column": [
        "NEd", "VEd", "Gk", "Qk", "Le", "fy", "grade", "section_class",
        "gamma_Q_used", "compression_result", "flex_buckling_result",
        "shear_result",
    ],
    "beam": [
        "MEd", "VEd", "Gk", "Qk", "q", "P", "L", "Mcr", "C1", "fy", "grade",
        "section_class", "gamma_Q_used", "bending_result", "ltb_result",
        "shear_result", "deflection_result",
    ],
}

QUICK_ANSWER_SYSTEM = """You are COMPLY, an engineering compliance assistant. \
Answer direct questions about values, sections, clauses, or fixes in flowing \
professional prose — full sentences, full words, no bullet lists, no tables, \
no headings. Include the specific value or answer with units. When suggesting \
a fix, two long comma-separated sentences is the target shape, quantifying the \
improvement (areas, radii of gyration, slenderness, reduction factors, \
resistances) rather than hand-waving. Cite code references in FULL ("Eurocode 3, \
Section 6.3.1"). NO em dashes. Use **bold** for key values and code references. \
If retrieved clause excerpts are provided, ground your quotes in them. If the \
question genuinely requires a fresh compliance audit, say so and suggest \
running the audit instead of guessing."""


# ---------------------------------------------------------------------------
# Classification / upload helpers (unchanged contracts)
# ---------------------------------------------------------------------------

async def classify_document(filename: str, text_sample: str) -> str:
    filename_lower = filename.lower()
    standard_filename_tokens = (
        "ec1", "ec2", "ec3", "ec4", "ec5", "ec6", "ec7", "ec8",
        "eurocode", "eur", "en 199", "en199", "bs ", "bs-", "qcs",
        "client_req", "clientreq", "requirements", "specification",
    )
    if any(token in filename_lower for token in standard_filename_tokens):
        return "STANDARD"

    project_filename_tokens = (
        "design", "calculation", "calc", "report", "solution",
        "coursework", "project", "submission",
    )
    if any(token in filename_lower for token in project_filename_tokens):
        return "PROJECT"

    try:
        prompt = (
            "Classify this engineering document as either STANDARD "
            "or PROJECT based on the filename and document sample.\n\n"
            "STANDARD documents — published codes, standards, and "
            "specifications (Eurocodes, EN 199x, British Standards, QCS, "
            "AASHTO, ACI, AISC, ISO) and client requirement documents.\n\n"
            "PROJECT documents — engineer-produced project-specific "
            "documents: design reports, calculation packs, geotechnical or "
            "site investigation reports, coursework submissions.\n\n"
            f"Filename: {filename}\n\n"
            "Document sample (first 500 characters):\n"
            f"{text_sample[:500]}\n\n"
            "Reply with exactly one word: STANDARD or PROJECT"
        )
        response = await anthropic_client.messages.create(
            model=ANTHROPIC_MODEL,
            max_tokens=10,
            system="Reply with exactly one word: STANDARD or PROJECT.",
            messages=[{"role": "user", "content": prompt}],
        )
        result = response.content[0].text.strip().upper()
        match = re.match(r"^(STANDARD|PROJECT)\b", result)
        if not match:
            raise ValueError(f"Unexpected classification response: {result!r}")
        return match.group(1)
    except Exception as exc:
        logger.error("classify_document error for %s: %s", filename, exc)
        return "UNKNOWN"


def _ext_to_type(ext: str) -> str:
    return "pdf" if ext == ".pdf" else "excel"


def _safe_filename(name: str) -> str:
    base = Path(name or "upload").name
    base = re.sub(r"[^\w.\- ()\[\]]", "_", base)
    return base[:120] or "upload"


def write_file_meta(file_id: str, updates: dict) -> dict:
    meta_path = UPLOAD_DIR / f"{file_id}_meta.json"
    meta = {}
    if meta_path.exists():
        try:
            meta = json.loads(meta_path.read_text())
        except Exception:
            meta = {}
    meta.update(updates)
    meta_path.write_text(json.dumps(meta))
    return meta


@asynccontextmanager
async def lifespan(app: FastAPI):
    if not ANTHROPIC_API_KEY:
        raise RuntimeError("ANTHROPIC_API_KEY is not configured")
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    AUDITS_DIR.mkdir(parents=True, exist_ok=True)
    mcp = MCPClient()
    await mcp.start()
    app.state.mcp = mcp
    logger.info("Startup complete — MCP connected: %s", mcp.connected)
    yield
    await app.state.mcp.stop()
    logger.info("Shutdown complete")


app = FastAPI(title="COMPLY API", lifespan=lifespan)

UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
app.mount("/uploads", StaticFiles(directory=str(UPLOAD_DIR)), name="uploads")

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_methods=["*"],
    allow_headers=["*"],
    allow_credentials=True,
)


class ChatMessage(BaseModel):
    role: Literal["user", "assistant"]
    content: str


class ChatRequest(BaseModel):
    message: str
    file_id: Optional[str] = None
    history: list[ChatMessage] = Field(default_factory=list)


class UploadResponse(BaseModel):
    file_id: str
    filename: str
    type: Literal["pdf", "excel"]
    classification: Literal["STANDARD", "PROJECT", "UNKNOWN"]
    rag_status: Optional[Literal["not_applicable", "pending", "embedded", "failed"]] = None
    rag_collection: Optional[str] = None
    rag_error: Optional[str] = None
    # Stage 0 result — populated only when a PROJECT pdf was just indexed
    indexed_elements: Optional[list[dict]] = None


class CellEdit(BaseModel):
    sheet: str
    row: int
    col: int
    value: str


# ---------------------------------------------------------------------------
# Standard endpoints
# ---------------------------------------------------------------------------

@app.get("/health")
async def health():
    mcp = app.state.mcp
    return {
        "status": "ok",
        "mcp_connected": mcp.connected,
        "mcp_pid": None,
        "tools": [t["name"] for t in mcp.get_tools()],
        "cache_size": {"coords": len(_coords_cache)},
    }


@app.post("/upload", response_model=UploadResponse)
async def upload(file: UploadFile = File(...)):
    original_name = _safe_filename(file.filename or "")
    ext = Path(original_name).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail="Unsupported file type")

    file_id = f"{uuid.uuid4()}_{original_name}"
    dest = UPLOAD_DIR / file_id

    contents = await file.read()
    dest.write_bytes(contents)
    logger.info("Uploaded: %s", file_id)

    file_type = _ext_to_type(ext)

    try:
        if file_type == "pdf":
            sample = await asyncio.to_thread(read_pdf_sample, str(dest), 3)
        else:
            sample = (await asyncio.to_thread(read_file, str(dest), file_type))[:2000]
    except Exception:
        sample = ""

    classification = await classify_document(original_name, sample)
    logger.info("Classified %s as %s", file_id, classification)

    write_file_meta(file_id, {
        "classification": classification,
        "filename": original_name,
        "type": file_type,
        "rag_status": "not_applicable",
        "rag_collection": None,
        "rag_error": None,
    })

    indexed_elements = None
    if file_type == "pdf":
        # Stage 0a — word coordinates for citation highlighting
        try:
            coords = await asyncio.to_thread(extract_pdf_coordinates, str(dest))
            (UPLOAD_DIR / f"{file_id}_coords.json").write_text(json.dumps(coords))
        except Exception as exc:
            logger.error("Coordinate extraction failed for %s: %s", file_id, exc)
        # Stage 0b — deterministic document index (elements, pages, designations)
        if classification == "PROJECT":
            try:
                index = await asyncio.to_thread(build_document_index, str(dest))
                (UPLOAD_DIR / f"{file_id}_index.json").write_text(json.dumps(index))
                indexed_elements = [
                    {"type": e["type"], "designation": e.get("designation")}
                    for e in index["elements"]
                ]
                logger.info("Indexed %s: %s", file_id,
                            [(e["type"], e.get("designation")) for e in index["elements"]])
            except Exception as exc:
                logger.error("Indexing failed for %s: %s", file_id, exc)

    return {
        "file_id": file_id,
        "filename": original_name,
        "type": file_type,
        "classification": classification,
        "rag_status": "not_applicable",
        "rag_collection": None,
        "rag_error": None,
        "indexed_elements": indexed_elements,
    }


@app.get("/files", response_model=list[UploadResponse])
async def list_files():
    if not UPLOAD_DIR.exists():
        return []
    result = []
    for f in UPLOAD_DIR.iterdir():
        if not f.is_file():
            continue
        if f.name.endswith(("_meta.json", "_coords.json", "_index.json")) or "_extract_" in f.name:
            continue
        ext = f.suffix.lower()
        parts = f.name.split("_", 1)
        original = parts[1] if len(parts) == 2 else f.name

        meta_path = UPLOAD_DIR / f"{f.name}_meta.json"
        classification = "PROJECT"
        rag_status = "not_applicable"
        if meta_path.exists():
            try:
                meta = json.loads(meta_path.read_text())
                classification = meta.get("classification", "PROJECT")
                rag_status = meta.get("rag_status", "not_applicable")
            except Exception:
                pass

        result.append({
            "file_id": f.name,
            "filename": original,
            "type": _ext_to_type(ext),
            "classification": classification,
            "rag_status": rag_status,
            "rag_collection": None,
            "rag_error": None,
        })
    return result


@app.delete("/files/{file_id}")
async def delete_file(file_id: str):
    path = UPLOAD_DIR / file_id
    if not path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    path.unlink()
    for suffix in ("_meta.json", "_coords.json", "_index.json",
                   "_extract_column.json", "_extract_beam.json"):
        side = UPLOAD_DIR / f"{file_id}{suffix}"
        if side.exists():
            side.unlink()
    _coords_cache.pop(file_id, None)
    logger.info("Deleted: %s", file_id)
    return {"deleted": True}


@app.get("/files/{file_id}/classification")
async def get_file_classification(file_id: str):
    meta_path = UPLOAD_DIR / f"{file_id}_meta.json"
    if not meta_path.exists():
        raise HTTPException(status_code=404, detail="Classification metadata not found")
    try:
        meta = json.loads(meta_path.read_text())
        return {"file_id": file_id, "classification": meta.get("classification", "PROJECT")}
    except Exception:
        raise HTTPException(status_code=500, detail="Failed to read classification metadata")


@app.get("/files/{file_id}/coordinates")
async def get_file_coordinates(file_id: str):
    if file_id in _coords_cache:
        return _coords_cache[file_id]
    coords_path = UPLOAD_DIR / f"{file_id}_coords.json"
    if not coords_path.exists():
        raise HTTPException(status_code=404, detail="Coordinates not found for this file")
    try:
        data = json.loads(coords_path.read_text())
        if len(_coords_cache) >= _COORDS_CACHE_MAX:
            _coords_cache.pop(next(iter(_coords_cache)))
        _coords_cache[file_id] = data
        return data
    except Exception:
        raise HTTPException(status_code=500, detail="Failed to read coordinates file")


@app.post("/admin/backfill-coordinates")
async def backfill_coordinates(x_admin_api_key: str | None = Header(default=None)):
    if ADMIN_API_KEY and x_admin_api_key != ADMIN_API_KEY:
        raise HTTPException(status_code=401, detail="Invalid admin API key")
    if not ADMIN_API_KEY:
        raise HTTPException(status_code=403, detail="Admin endpoint disabled")
    if not UPLOAD_DIR.exists():
        return {"backfilled": 0, "skipped": 0, "errors": []}
    backfilled, skipped, errors = 0, 0, []
    for f in UPLOAD_DIR.iterdir():
        if not f.is_file() or f.suffix.lower() != ".pdf":
            continue
        if f.name.endswith(("_meta.json", "_coords.json", "_index.json")) or "_extract_" in f.name:
            continue
        coords_path = UPLOAD_DIR / f"{f.name}_coords.json"
        if coords_path.exists():
            skipped += 1
            continue
        try:
            coords = await asyncio.to_thread(extract_pdf_coordinates, str(f))
            coords_path.write_text(json.dumps(coords))
            backfilled += 1
        except Exception as exc:
            errors.append({"file": f.name, "error": str(exc)})
    return {"backfilled": backfilled, "skipped": skipped, "errors": errors}


@app.patch("/files/{file_id}/cell")
async def edit_cell(file_id: str, request: CellEdit):
    path = UPLOAD_DIR / file_id
    if not path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    ext = path.suffix.lower()
    if ext not in {".xlsx", ".xls"}:
        raise HTTPException(status_code=400, detail="File is not an Excel file")
    wb = load_workbook(path)
    if request.sheet not in wb.sheetnames:
        raise HTTPException(status_code=404, detail="Sheet not found")
    ws = wb[request.sheet]
    ws.cell(row=request.row + 2, column=request.col + 1).value = request.value
    wb.save(path)
    return {"saved": True, "sheet": request.sheet, "row": request.row,
            "col": request.col, "value": request.value}


# ---------------------------------------------------------------------------
# Chat: intent routing + the deterministic audit pipeline + quick answers
# ---------------------------------------------------------------------------

def _sse(payload: dict) -> str:
    return f"data: {json.dumps(payload)}\n\n"


def is_compliance_audit(message: str) -> bool:
    msg = message.strip()
    if msg.startswith(COMPLIANCE_AUDIT_PREFIX):
        return True
    verbs = re.search(r"\b(check|audit|flag|verify|assess|re-?run)\b", msg, re.I)
    targets = re.search(r"\b(column|beam|compliance|violation|design|element)s?\b", msg, re.I)
    return bool(verbs and targets)


def clean_message(message: str) -> str:
    msg = message.strip()
    if msg.startswith(COMPLIANCE_AUDIT_PREFIX):
        msg = msg[len(COMPLIANCE_AUDIT_PREFIX):].strip()
    return msg


_STATUS_WORDS = {
    "failure": "FAIL", "fail": "FAIL", "error": "ERROR", "conflict": "CONFLICT",
    "missing": "MISSING", "assumed": "ASSUMED", "warning": "WARNING",
    "pass": "PASS",
}


def _scope_filters(message: str) -> tuple[str | None, set[str]]:
    """Spec Part 10: the prompt controls SCOPE. 'clause named -> scoped to that
    clause; status named -> filtered by that status'."""
    msg = message.lower()
    clause = None
    m = re.search(r"\b(?:section|clause|§)\s*(\d+(?:\.\d+){0,3})", msg)
    if m:
        clause = m.group(1)
    statuses = {v for k, v in _STATUS_WORDS.items() if re.search(rf"\b{k}(?:s|es|ures)?\b", msg)}
    # plain "check/audit" verbs shouldn't trigger the PASS/FAIL words heuristic
    if statuses == {"PASS"} and "pass" in msg and "compliance" in msg:
        statuses = set()
    return clause, statuses


def _apply_scope_filters(findings: list[dict], clause: str | None,
                         statuses: set[str]) -> list[dict]:
    out = findings
    if clause:
        filtered = [f for f in out if clause in (f.get("clause") or "")]
        if filtered:
            out = filtered
    if statuses:
        filtered = [f for f in out if f.get("status") in statuses]
        if filtered:
            out = filtered
    return out


def _scope_elements(message: str, indexed: list[dict]) -> list[dict]:
    msg = message.lower()
    wants_col = "column" in msg
    wants_beam = "beam" in msg
    if wants_col and not wants_beam:
        chosen = [e for e in indexed if e["type"] == "column"]
    elif wants_beam and not wants_col:
        chosen = [e for e in indexed if e["type"] == "beam"]
    else:
        chosen = indexed
    return chosen or indexed


def _resolve_project_pdf(file_id: str | None) -> Path | None:
    """The audit target: the given file if it's a PROJECT pdf, else the most
    recently uploaded PROJECT pdf."""
    def is_project_pdf(p: Path) -> bool:
        if p.suffix.lower() != ".pdf" or p.name.endswith(
                ("_meta.json", "_coords.json", "_index.json")):
            return False
        meta_path = UPLOAD_DIR / f"{p.name}_meta.json"
        try:
            meta = json.loads(meta_path.read_text())
            return meta.get("classification") == "PROJECT"
        except Exception:
            return False

    if file_id:
        p = UPLOAD_DIR / file_id
        if p.exists() and is_project_pdf(p):
            return p
    candidates = [p for p in UPLOAD_DIR.iterdir() if p.is_file() and is_project_pdf(p)]
    if not candidates:
        return None
    return max(candidates, key=lambda p: p.stat().st_mtime)


async def _load_index(pdf_path: Path) -> dict:
    index_path = UPLOAD_DIR / f"{pdf_path.name}_index.json"
    if index_path.exists():
        try:
            return json.loads(index_path.read_text())
        except Exception:
            pass
    index = await asyncio.to_thread(build_document_index, str(pdf_path))
    index_path.write_text(json.dumps(index))
    return index


async def _call_tool_json(mcp, name: str, args: dict, timeout: float = 120) -> dict:
    result = await mcp.call_tool(name, args, timeout=timeout)
    first = result.content[0] if result.content else None
    text = getattr(first, "text", None)
    if not text:
        raise RuntimeError(f"{name} returned no content")
    return json.loads(text)


def _persist_audit(document: str, payload: dict) -> str:
    ts = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    stem = re.sub(r"[^\w-]", "_", Path(document).stem)[:60]
    path = AUDITS_DIR / f"{stem}-audit-{ts}.json"
    path.write_text(json.dumps(payload, indent=2, default=str))
    # Best-effort Firestore history (spec 5.1 audits/); never blocks the audit.
    try:
        from google.cloud import firestore
        from google.oauth2 import service_account
        creds_path = os.getenv("GOOGLE_APPLICATION_CREDENTIALS",
                               str(Path(__file__).parent / "firebase-key.json"))
        creds = service_account.Credentials.from_service_account_file(creds_path)
        db = firestore.Client(project=os.getenv("FIREBASE_PROJECT_ID"), credentials=creds)
        db.collection("audits").document(path.stem).set(
            json.loads(json.dumps(payload, default=str)))
    except Exception as exc:
        logger.warning("Firestore audit persist skipped: %s", exc)
    return str(path)


async def _audit_stream(req: ChatRequest):
    mcp = app.state.mcp
    timing: dict[str, float] = {}
    message = clean_message(req.message)

    pdf_path = _resolve_project_pdf(req.file_id)
    if pdf_path is None:
        yield _sse({"type": "error",
                    "message": "No PROJECT design report uploaded — upload the design PDF first."})
        yield _sse({"type": "done"})
        return
    document = pdf_path.name.split("_", 1)[-1]

    # Stage 0/2 — index + identify (deterministic)
    t0 = time.perf_counter()
    yield _sse({"type": "status", "message": "Indexing document…"})
    index = await _load_index(pdf_path)
    elements = _scope_elements(message, index.get("elements") or [])
    timing["index_s"] = round(time.perf_counter() - t0, 2)
    if not elements:
        yield _sse({"type": "error",
                    "message": "No structural elements (column/beam) found in the document index."})
        yield _sse({"type": "done"})
        return

    # Stage 3/4 — cache check, then extract (AI call #1, per element).
    # The extraction cache is keyed on the element's pages, so a re-run of an
    # unchanged document skips the AI call entirely (spec Stage 3).
    extracts: dict[str, dict] = {}
    t1 = time.perf_counter()
    for el in elements:
        el_type = el["type"]
        cache_path = UPLOAD_DIR / f"{pdf_path.name}_extract_{el_type}.json"
        cache_key = json.dumps(el.get("pages") or [])
        if cache_path.exists():
            try:
                cached = json.loads(cache_path.read_text())
                if cached.get("_pages_key") == cache_key:
                    yield _sse({"type": "status",
                                "message": f"Using cached extraction — {el_type}…"})
                    extracts[el_type] = cached["extract"]
                    continue
            except Exception:
                pass
        yield _sse({"type": "status",
                    "message": f"Extracting design values — {el_type}…"})
        try:
            required = REQUIRED_EXTRACT_INPUTS[el_type]
            extracted = await _call_tool_json(mcp, "extract_element", {
                "element_type": el_type,
                "pdf_path": str(pdf_path),
                "pages": el.get("pages") or [],
                "required_inputs": required,
                "input_descriptions": {k: INPUT_DESCRIPTIONS[k] for k in required},
                "designation": el.get("designation"),
            })
        except Exception as exc:
            logger.error("extract_element failed (%s): %s", el_type, exc)
            yield _sse({"type": "status",
                        "message": f"Extraction failed for {el_type} — treating its inputs as missing…"})
            extracted = {"element": el_type, "designation": el.get("designation"),
                         "values": {}, "not_found": REQUIRED_EXTRACT_INPUTS[el_type]}
        extracted["document"] = document
        extracts[el_type] = extracted
        try:
            cache_path.write_text(json.dumps(
                {"_pages_key": cache_key, "extract": extracted}))
        except Exception:
            pass
    timing["extract_s"] = round(time.perf_counter() - t1, 2)

    # Stage 5 — deterministic engine (pure Python)
    yield _sse({"type": "status", "message": "Running deterministic engine…"})
    t2 = time.perf_counter()
    all_findings: list[dict] = []
    designations: dict[str, str | None] = {}
    for el_type, extracted in extracts.items():
        designation = extracted.get("designation")
        designation = normalize_designation(designation) if designation else None
        designations[el_type] = designation
        props = get_section(designation) if designation else None
        try:
            findings = await asyncio.to_thread(run_audit, el_type, extracted, props)
        except Exception as exc:
            logger.exception("engine failed for %s", el_type)
            yield _sse({"type": "status",
                        "message": f"Engine error on {el_type} ({exc}) — element skipped…"})
            continue
        for f in findings:
            f["element"] = el_type
            f["designation"] = designation
        all_findings.extend(findings)
    # CONFLICT detection runs across every element of the document, pulling
    # cached extracts for elements outside the current scope so a single-
    # element audit still catches cross-element disagreements (spec 14.4).
    conflict_inputs = dict(extracts)
    for el in index.get("elements") or []:
        if el["type"] in conflict_inputs:
            continue
        side = UPLOAD_DIR / f"{pdf_path.name}_extract_{el['type']}.json"
        if side.exists():
            try:
                conflict_inputs[el["type"]] = json.loads(side.read_text())["extract"]
            except Exception:
                pass
    if len(conflict_inputs) > 1:
        for f in cross_reference(conflict_inputs):
            f["element"] = "all"
            all_findings.append(f)

    clause_filter, status_filter = _scope_filters(message)
    all_findings = _apply_scope_filters(all_findings, clause_filter, status_filter)
    order = ["FAIL", "ERROR", "CONFLICT", "MISSING", "ASSUMED", "WARNING", "PASS"]
    all_findings.sort(key=lambda x: order.index(x["status"]))
    pills = summarize(all_findings)
    timing["engine_s"] = round(time.perf_counter() - t2, 3)

    if len(elements) == 1:
        el = elements[0]
        title = f"{el['type'].capitalize()} compliance check"
        subtitle_parts = [document]
        if designations.get(el["type"]):
            subtitle_parts.append(designations[el["type"]])
    else:
        title = "Full compliance audit"
        subtitle_parts = [document,
                          ", ".join(e["type"] for e in elements)]
    subtitle = " . ".join(subtitle_parts)

    findings_payload = {
        "title": title,
        "subtitle": subtitle,
        "document": document,
        "file_id": pdf_path.name,
        "pills": pills,
        "findings": all_findings,
    }
    yield _sse({"type": "findings", "data": findings_payload})

    # Stage 6 — overview (AI call #2, narration only)
    yield _sse({"type": "status", "message": "Writing overview…"})
    t3 = time.perf_counter()
    primary = elements[0]["type"]
    try:
        overview = await _call_tool_json(mcp, "write_overview", {
            "element": primary if len(elements) == 1 else "all elements",
            "designation": designations.get(primary),
            "document": document,
            "findings": [
                {k: f.get(k) for k in
                 ("status", "name", "clause", "issue", "metrics", "element")}
                for f in all_findings
            ],
            "summary": pills,
        }, timeout=90)
    except Exception as exc:
        logger.error("write_overview failed: %s", exc)
        overview = {
            "overview": "The overview writer was unavailable; the findings table above is complete and authoritative.",
            "recommended_actions": [],
        }
    timing["overview_s"] = round(time.perf_counter() - t3, 2)
    timing["total_s"] = round(time.perf_counter() - t0, 2)

    yield _sse({"type": "overview", "data": overview})

    audit_payload = {**findings_payload, **overview, "timing": timing,
                     "created_at": datetime.now(timezone.utc).isoformat()}
    try:
        path = await asyncio.to_thread(_persist_audit, document, audit_payload)
        logger.info("Audit persisted: %s (timing %s)", path, timing)
    except Exception as exc:
        logger.warning("Audit persistence failed: %s", exc)

    yield _sse({"type": "done", "timing": timing})


async def _quick_answer_stream(req: ChatRequest):
    mcp = app.state.mcp
    message = clean_message(req.message)

    yield _sse({"type": "status", "message": "Searching the standards…"})
    refs: list[dict] = []
    try:
        search = await _call_tool_json(mcp, "search_documents",
                                       {"query": message, "top_k": 3}, timeout=45)
        refs = search.get("results") or []
    except Exception as exc:
        logger.warning("search_documents failed: %s", exc)

    if refs:
        yield _sse({"type": "quick_refs", "data": [
            {"clause": r.get("clause"), "source": r.get("source"),
             "page": r.get("page")} for r in refs
        ]})

    history = [
        {"role": m.role, "content": m.content[:MAX_HISTORY_CHARS]}
        for m in req.history[-MAX_HISTORY_MESSAGES:]
    ]
    context = ""
    if refs:
        context = ("\n\nRetrieved clause excerpts (ground your citations in "
                   "these):\n" + json.dumps(refs))
    messages = history + [{"role": "user", "content": message + context}]

    first_token = False
    for attempt in range(4):
        try:
            async with anthropic_client.messages.stream(
                model=ANTHROPIC_MODEL,
                max_tokens=1024,
                system=[{"type": "text", "text": QUICK_ANSWER_SYSTEM,
                         "cache_control": {"type": "ephemeral"}}],
                messages=messages,
            ) as stream:
                async for text in stream.text_stream:
                    first_token = True
                    yield _sse({"type": "token", "content": text})
            break
        except (anthropic.RateLimitError, anthropic.APIStatusError,
                anthropic.APIConnectionError) as exc:
            status = getattr(exc, "status_code", None)
            retryable = isinstance(
                exc, (anthropic.RateLimitError, anthropic.APIConnectionError)
            ) or status in (429, 500, 503, 529)
            if first_token or not retryable or attempt == 3:
                yield _sse({"type": "error", "message": f"Model unavailable: {exc}"})
                break
            await asyncio.sleep(2 ** attempt)
    yield _sse({"type": "done"})


@app.post("/chat")
async def chat(req: ChatRequest):
    async def generate():
        async with _request_lock:
            try:
                if is_compliance_audit(req.message):
                    async for event in _audit_stream(req):
                        yield event
                else:
                    async for event in _quick_answer_stream(req):
                        yield event
            except Exception as exc:
                logger.exception("chat pipeline error")
                yield _sse({"type": "error", "message": str(exc)})
                yield _sse({"type": "done"})

    return StreamingResponse(generate(), media_type="text/event-stream")
