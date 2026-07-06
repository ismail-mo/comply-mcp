import logging
import pdfplumber
import re
from pathlib import Path
import pypdf
import openpyxl

logger = logging.getLogger("comply.reader")


def read_pdf_sample(file_path: str, max_pages: int = 3) -> str:
    """Extract first max_pages pages with layout-aware spacing for classification."""
    path = Path(file_path)
    try:
        reader = pypdf.PdfReader(str(path))
        pages = []
        for i, page in enumerate(reader.pages):
            if i >= max_pages:
                break
            try:
                text = page.extract_text(extraction_mode="layout")
            except TypeError:
                text = page.extract_text()
            if not text:
                continue
            # Collapse runs of spaces/tabs to a single space
            text = re.sub(r'[ \t]+', ' ', text)
            # Collapse excessive blank lines
            text = re.sub(r'\n{3,}', '\n\n', text)
            pages.append(f"[PAGE {i + 1}]\n{text.strip()}")
        logger.info("PDF sample read: %s (%d pages)", path.name, len(pages))
        return "\n\n".join(pages)
    except Exception as exc:
        raise ValueError(f"Could not read PDF sample '{path}': {exc}") from exc


def extract_pdf_coordinates(file_path: str) -> dict:
    path = Path(file_path)
    try:
        result = {"page_count": 0, "pages": {}}
        with pdfplumber.open(str(path)) as pdf:
            result["page_count"] = len(pdf.pages)
            for i, page in enumerate(pdf.pages, start=1):
                raw_words = page.extract_words()
                words = [
                    {
                        "text": w["text"],
                        "x0": round(w["x0"], 2),
                        "y0": round(w["top"], 2),
                        "x1": round(w["x1"], 2),
                        "y1": round(w["bottom"], 2),
                    }
                    for w in raw_words
                ]
                result["pages"][str(i)] = {
                    "width": round(page.width, 2),
                    "height": round(page.height, 2),
                    "words": words,
                }
        logger.info("PDF coordinates extracted: %s (%d pages)", path.name, result["page_count"])
        return result
    except Exception as exc:
        logger.error("Failed to extract PDF coordinates for '%s': %s", path.name, exc)
        return {"page_count": 0, "pages": {}}


def read_excel(file_path: str) -> str:
    path = Path(file_path)
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        sheets = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines = [f"=== Sheet: {sheet_name} ==="]
            for row in ws.iter_rows():
                values = ["" if cell.value is None else str(cell.value) for cell in row]
                if any(v != "" for v in values):
                    lines.append(" | ".join(values))
            sheets.append("\n".join(lines))
        wb.close()
        logger.info("Excel read: %s (%d sheets)", path.name, len(sheets))
        return "\n\n".join(sheets)
    except Exception as exc:
        raise ValueError(f"Could not read Excel file '{path}': {exc}") from exc


def read_file(file_path: str, file_type: str) -> str:
    # Whole-PDF reads were removed deliberately: PDF content only enters the
    # system through chunked RAG retrieval (services/rag.py -> project_chunks).
    if file_type == "excel":
        return read_excel(file_path)
    raise ValueError(
        f"Unsupported file_type '{file_type}': whole-document PDF reads are not allowed"
    )
